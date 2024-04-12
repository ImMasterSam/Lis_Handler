from os import listdir
from openpyxl import Workbook, load_workbook

excel_path = "Data.xlsx"

try:
    ex = load_workbook(excel_path)
except:
    print(f"ERROR: 無法載入 Ecxel 檔 {excel_path}")
    exit()



def load_Delay(name, file, start):

    count = 0
    Delay_type = {"tphl" : 4, "tplh": 1, "tr": 7, "tf" : 8}

    try:
        ws = ex[name]
    except:
        return -1
    
    for i in range(len(file)):

        s = file[i].strip()

        #先抓開頭
        if s == "****** transient analysis tnom=  25.000 temp=  25.000 ******":
            group = file[i:i+8]

            #如果第二行不是則跳出
            if group[1].strip().endswith("."):
                continue
            
            #判斷 cl 值
            cl = 0
            if group[1].find("parameter cl") != -1:
                cl = float(group[1][22:22+13]) // 10
            else:
                cl = 0

            #讀取各行資料並輸入至Excel
            for lines in group:
                for T in Delay_type.keys():

                    if lines[1:len(T)+1] == T:
                        data = float(lines[len(T)+2: len(T)+2+15]) / 1000
                        if lines[len(T)+2: len(T)+17] == 'p':
                            data /= 1000
                        ws[f"{chr(ord(start[0]) + Delay_type[T])}{int(start[1:]) + 5 + int(cl)}"] = data
                        count += 1

    if count == 16:
        return 1
    else:
        return -2

def load_Power(name, file, start, cl):

    count = 0

    try:
        ws = ex[name]
    except:
        return -1

    for i in range(len(file)):
        
        if file[i].find("meas_variable = min_power") != -1:
            line = file[i+2].strip()
            idx = line.find("avgdev")
            data = float(line[idx+9:idx+9+14])
            if line[idx+9+14] == 'p':
                data /= 1000
            ws[f"{chr(ord(start[0]) + 9)}{int(start[1:]) + 5 + int(cl)}"] = data
            count += 1


        if file[i].find("meas_variable = src_power") != -1:
            line = file[i+2].strip()
            idx = line.find("avgdev")
            data = float(line[idx+9:idx+9+14])
            if line[idx+9+14] == 'p':
                data /= 1000
            ws[f"{chr(ord(start[0]) + 10)}{int(start[1:]) + 5 + int(cl)}"] = data
            count += 1

    if count == 2:
        return 1
    else:
        return -2

for names in listdir():
    
    if names.endswith(".lis"):
        
        feature = list(map(str, names[:-4].split("_")))
        try:
            Pre_Post, Name, scale, Type, cl = feature[:5]
        except:
            print(f"ERROR -> {names}: 檔名格式錯誤")
            continue
        Name = Name + "_" + scale

        input = ""
        if len(feature) > 5:
            input = feature[5]

        #設定儲存格起點
        start = ""
        if Pre_Post == "Pre":
            start += "A"
        elif Pre_Post == "Post":
            start += "N"
        if input == "B":
            start += "12"
        else:
            start += "1"
        
        res = 0
        file = open(names, "r").readlines()
        if Type == 'T' :
            res = load_Delay(Name, file, start)
        elif Type == 'P':
            res = load_Power(Name, file, start, int(cl)//10)

        if res == -1:
            print(f"ERROR -> {names}: 工作表 {Name} 不存在")
        elif res == -2:
            print(f"ERROR -> {names}: 無法完整加載所有數據，請檢查 .lis 檔")

#ex.save(excel_path)